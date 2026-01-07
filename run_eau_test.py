import os
import glob
import pandas as pd
import numpy as np
import shutil
from functools import partial

# ==========================================
# [关键步骤] 1. 配置 XLA Dump 环境变量
# ==========================================
# 必须在 import jax 或调用 jax 相关函数之前设置
# 这样 XLA 编译器初始化时才会读取这些 Flag

# 定义 dump 根目录
XLA_DUMP_DIR = "xla_dump_debug"
XLA_JF_DUMP_DIR = "xla_jf_dump_debug"

# 清理旧的 dump 目录以防混淆
# if os.path.exists(XLA_DUMP_DIR):
#     shutil.rmtree(XLA_DUMP_DIR)
# os.makedirs(XLA_DUMP_DIR)

if os.path.exists(XLA_JF_DUMP_DIR):
    shutil.rmtree(XLA_JF_DUMP_DIR)
os.makedirs(XLA_JF_DUMP_DIR)

# 设置 XLA FLAGS
# --xla_dump_to: 输出目录
# --xla_dump_hlo_as_text: 输出文本格式
# --xla_dump_hlo_as_dot: 输出 GraphViz dot 格式
# --xla_dump_hlo_pass_re: 正则表达式，过滤只输出关键步骤 (如 '.*' 输出所有步骤，或 'optimized' 只输出优化后)
# os.environ['XLA_FLAGS'] = (
#     f"--xla_dump_to={XLA_DUMP_DIR} "
#     "--xla_dump_hlo_as_text "
#     "--xla_dump_hlo_as_dot "
# )

os.environ['LIBTPU_INIT_ARGS'] = f"--xla_jf_dump_to={XLA_JF_DUMP_DIR}"
print(f"Set LIBTPU_INIT_ARGS to: {os.environ['LIBTPU_INIT_ARGS']}")


import jax
import jax.numpy as jnp
from jax import lax

# 引入 Excel 样式库
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 设置 JAX 平台
# os.environ['JAX_PLATFORM_NAME'] = 'tpu'

def _exp2_wrapper(x, accuracy_mode=None):
  return lax.exp2_p.bind(x, accuracy=accuracy_mode)    # None, lax.AccuracyMode.DEFAULT, lax.AccuracyMode.HIGHEST


def export_exp2_hlo_variants_tpu():
    """
    针对 TPU 触发编译，生成 .txt 和 .dot 到 xla_dump_debug 文件夹
    """
    print(f"\n--- Triggering XLA Compilation for exp2 (Check folder: {XLA_DUMP_DIR}) ---")

    # None, lax.AccuracyMode.DEFAULT, lax.AccuracyMode.HIGHEST
    modes = [None]
    
    # 构造 Dummy Input
    dummy_input = jnp.array([1.5, -2.0, 0.0], dtype=jnp.float32)
    
    for mode in modes:
        print(f"  > Compiling mode: {mode} ...")
        
        # 1. 构造特定 mode 的函数
        func = partial(_exp2_wrapper, accuracy_mode=mode)
        jit_func = jax.jit(func)
        
        try:
            # 2. Lowering
            lowered = jit_func.lower(dummy_input)
            
            # 3. Compile (关键！)
            # 只有调用 compile()，XLA 才会执行优化管道并根据环境变量 dump 文件
            compiled = lowered.compile()

        except Exception as e:
            print(f"  [Error] Compilation failed for {mode}: {e}")

    print(f"  -> All dumps saved to ./{XLA_DUMP_DIR}/")
    print("     You will see files like 'module_000x.optimized_module.dot'")
    print("----------------------------------------------------\n")


# ==========================================
# 1. 算子映射表
# ==========================================
OP_MAP = {
    'exp2': _exp2_wrapper,
    "log2": lambda x: lax.div(lax.log(x), lax.log(2.0)),
    "rcp": lax.reciprocal,
    "rsqrt": lax.rsqrt,
    "sin": lax.sin,
    "cos": lax.cos,
    "tanh": lax.tanh,
    "erf": lax.erf,
    "sigshft": lambda x: 0.5 * lax.tanh(x / 2.0),
}


# ==========================================
# 2. 核心工具
# ==========================================
def hex_str_to_f32(hex_series):
    """Hex 字符串列 -> JAX Float32"""
    clean_hex = hex_series.astype(str).str.strip().str.replace(' ', '')
    int_values = []
    for h in clean_hex:
        try:
            if h.lower() == 'nan' or h == '':
                int_values.append(0)
            else:
                int_values.append(int(h, 16))
        except ValueError:
            int_values.append(0)

    np_uint32 = np.array(int_values, dtype=np.uint32)
    return lax.bitcast_convert_type(jnp.array(np_uint32), jnp.float32)


def f32_to_hex_str(jax_array):
    """JAX Float32 -> Hex 字符串列表"""
    val_uint32 = lax.bitcast_convert_type(jax_array, jnp.uint32)
    val_uint32.block_until_ready()
    np_vals = np.array(val_uint32)
    return [f"{x:08X}" for x in np_vals]


def get_op_func(filename):
    fname = os.path.basename(filename).lower()
    for key, func in OP_MAP.items():
        if key in fname: return func
    return None


# ==========================================
# 3. 处理逻辑 (含 Excel 美化)
# ==========================================
def process_file(filepath, output_dir):
    filename = os.path.basename(filepath)
    base_name = os.path.splitext(filename)[0]
    # 修改输出后缀为 .xlsx 以支持标红
    save_path = os.path.join(output_dir, f"RES_{base_name}.xlsx")

    print(f"Processing {filename}...")

    try:
        # 读取 CSV (全部读为字符串以保留格式)
        df = pd.read_csv(filepath, dtype=str)
    except Exception as e:
        print(f"  [Error] Read failed: {e}")
        return

    op_func = get_op_func(filename)
    if op_func is None:
        print(f"  [Skip] No operator found for {filename}")
        return

    if 'Xin' not in df.columns:
        print("  [Error] Column 'Xin' not found.")
        return

    # --- 1. 执行计算 ---
    jit_op = jax.jit(op_func)
    try:
        input_tensor = hex_str_to_f32(df['Xin'])
        output_tensor = jit_op(input_tensor)
        output_hex_list = f32_to_hex_str(output_tensor)
        df['TPUOut'] = output_hex_list
    except Exception as e:
        print(f"  [Error] Execution failed: {e}")
        return

    # --- 2. 对比结果 ---
    results = []
    if 'CmodelOut' in df.columns:
        for idx, row in df.iterrows():
            c_out = str(row['CmodelOut']).strip().upper()
            t_out = str(row['TPUOut']).strip().upper()

            # 简单的字符串对比 (忽略 NaN)
            if c_out == 'NAN' or c_out == '':
                results.append("")
            elif c_out == t_out:
                results.append("PASS")
            else:
                results.append(f"FAIL (Exp: {c_out})")
    else:
        results = [""] * len(df)

    df['Result'] = results

    # --- 3. 保存并美化 Excel ---
    try:
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Result')

            # 获取 worksheet 对象
            workbook = writer.book
            sheet = writer.sheets['Result']

            # 定义样式
            red_font = Font(color="FF0000", bold=True)
            alignment = Alignment(horizontal='left')

            # 调整列宽 (设置为 25)
            for col_idx in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = 25

            # 找到 'Result' 列的索引 (1-based)
            # 注意: df.to_excel 可能会重排，这里按顺序找
            res_col_idx = df.columns.get_loc('Result') + 1

            # 遍历数据行 (跳过第1行表头)
            for i in range(len(df)):
                row_idx = i + 2  # Excel row index (1-based, +1 header)

                # 获取 Result 单元格的值
                res_val = sheet.cell(row=row_idx, column=res_col_idx).value
                res_str = str(res_val) if res_val else ""

                # 如果是 FAIL，标红整行
                if res_str.startswith("FAIL"):
                    for col in range(1, len(df.columns) + 1):
                        cell = sheet.cell(row=row_idx, column=col)
                        cell.font = red_font

                # 可选: 统一左对齐
                for col in range(1, len(df.columns) + 1):
                    sheet.cell(row=row_idx, column=col).alignment = alignment

        # 打印统计
        pass_count = results.count("PASS")
        total_count = len([r for r in results if r])  # 只统计非空结果
        print(f"  -> Saved {save_path} (Pass: {pass_count}/{total_count})")

    except Exception as e:
        print(f"  [Error] Save Excel failed: {e}")


def main():
    print("=" * 60)
    devices = jax.devices()
    print(f"Running on: {devices}")
    print("=" * 60 + "\n")

    input_dir = "input_eau_260105"
    output_dir = "result_eau_260105"

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    files = glob.glob(os.path.join(input_dir, "*.csv"))
    print(f"Found {len(files)} files.")

    for f in files:
        process_file(f, output_dir)

    # 导出 XLA Dumps (txt + dot)
    export_exp2_hlo_variants_tpu()

if __name__ == "__main__":
    main()
