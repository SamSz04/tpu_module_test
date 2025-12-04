import os
import pandas as pd
import jax
import jax.numpy as jnp
from jax import lax
import logging

# 引入 Excel 样式库 (与 auto_tester 保持一致)
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 设置 JAX 运行平台
os.environ['JAX_PLATFORM_NAME'] = 'tpu'

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger()


# ==========================================
# 1. 核心工具函数 (Hex <-> JAX)
# ==========================================
def hex_to_jax(hex_str, dtype):
    """Hex 字符串转 JAX Array"""
    clean_hex = str(hex_str).strip().lower()
    if pd.isna(hex_str) or not clean_hex or clean_hex == 'nan':
        return jnp.array(0, dtype=dtype)
    if clean_hex.startswith('0x'): clean_hex = clean_hex[2:]
    clean_hex = clean_hex.replace(' ', '')

    try:
        int_val = int(clean_hex, 16)
    except ValueError:
        return jnp.array(0, dtype=dtype)

    # Float Types (Bitcast)
    if dtype == jnp.float32:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint32), jnp.float32)
    elif dtype == jnp.bfloat16:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint16), jnp.bfloat16)

    # Integer Types (Direct)
    return jnp.array(int_val, dtype=dtype)


def jax_to_hex(jax_val):
    """JAX Array 转 Hex 字符串"""
    dtype = jax_val.dtype
    width = 8  # default 32-bit
    val_uint = jax_val

    if dtype == jnp.float32:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint32)
        width = 8
    elif dtype == jnp.bfloat16:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint16)
        width = 4
    elif dtype in [jnp.int32, jnp.uint32]:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint32)
        width = 8
    elif dtype in [jnp.int16, jnp.uint16]:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint16)
        width = 4

    try:
        int_val = int(val_uint)
    except:
        int_val = int(val_uint.item())

    return f"0x{int_val:0{width}X}"


# ==========================================
# 2. 智能解析器 (复用 auto_tester 逻辑)
# ==========================================
def parse_test_file(filepath):
    """解析 Excel/CSV 中的 Definition 表和 Test 表"""
    try:
        # 支持 xlsx 直接读取，保留原始结构
        if filepath.endswith('.xlsx'):
            df_raw = pd.read_excel(filepath, header=None)
        else:
            df_raw = pd.read_csv(filepath, header=None, engine='python')
    except Exception as e:
        logger.error(f"  [Error] Read failed: {e}")
        return None, None

    def_start_row = -1
    test_start_row = -1

    # 扫描表头位置
    for i, row in df_raw.iterrows():
        row_str = " ".join([str(x).lower() for x in row.values])
        if 'type' in row_str and 'val' in row_str:
            def_start_row = i
        if ('test' in row_str or 'x' in row_str) and 'exp_out' in row_str:
            test_start_row = i

    if def_start_row == -1 or test_start_row == -1:
        logger.warning(f"  [Skip] Headers not found in {os.path.basename(filepath)}")
        return None, None

    # --- 提取 Definition Table ---
    header_row_def = df_raw.iloc[def_start_row]
    valid_cols_def = header_row_def.dropna().index
    df_def = df_raw.iloc[def_start_row + 1: test_start_row, valid_cols_def].copy()
    df_def.columns = header_row_def[valid_cols_def].astype(str).str.strip()

    # 过滤 "test" 杂行
    if not df_def.empty:
        is_junk = df_def.astype(str).apply(lambda x: x.str.contains('test', case=False)).any(axis=1)
        df_def = df_def[~is_junk]
    df_def = df_def.dropna(how='all')

    # --- 提取 Test Table ---
    header_row_test = df_raw.iloc[test_start_row]
    valid_cols_test = header_row_test.dropna().index
    df_test = df_raw.iloc[test_start_row + 1:, valid_cols_test].copy()
    df_test.columns = header_row_test[valid_cols_test].astype(str).str.strip()
    df_test = df_test.dropna(how='all')

    return df_def, df_test


# ==========================================
# 3. 加法测试类
# ==========================================
class TPUAddTester:
    def __init__(self, mode='fp32'):
        self.mode = mode
        # 确定 JAX 数据类型
        if mode == 'bf16':
            self.dtype = jnp.bfloat16
        else:
            self.dtype = jnp.float32

        print(f"[{mode.upper()}] Tester Initialized. Dtype: {self.dtype}")
        self.check_tpu()

    def check_tpu(self):
        devices = jax.devices()
        print(f"Running on devices: {devices}")
        if any('TPU' in str(d) for d in devices):
            print("✅ TPU Detected.")
        else:
            print("⚠️ WARNING: Running on CPU/GPU.")

    @staticmethod
    @jax.jit
    def run_kernel(x, y):
        """核心算子: 加法"""
        return x + y

    def run_test_file(self, input_file, output_file):
        if not os.path.exists(input_file):
            print(f"File not found: {input_file}")
            return

        print(f"\nProcessing {input_file}...")

        # 1. 解析文件
        df_def, df_test = parse_test_file(input_file)
        if df_def is None: return

        # 2. 构建 Value Registry (动态读取)
        value_registry = {}
        type_col = next((c for c in df_def.columns if 'type' in c.lower()), None)
        val_col = next((c for c in df_def.columns if 'val' in c.lower()), None)

        if type_col and val_col:
            for _, row in df_def.iterrows():
                if pd.notna(row[type_col]) and pd.notna(row[val_col]):
                    name = str(row[type_col]).strip()
                    hex_val = str(row[val_col]).strip()
                    # 使用当前模式的 dtype 进行转换
                    value_registry[name] = hex_to_jax(hex_val, self.dtype)
            print(f"  -> Loaded {len(value_registry)} definitions.")
        else:
            print("  -> ERROR: Definition table missing 'type' or 'val' columns.")
            return

        # 3. 运行测试
        test_outs = []
        results = []

        # 识别列名
        x_col = next((c for c in df_test.columns if c.strip() == 'x'), None)
        y_col = next((c for c in df_test.columns if c.strip() == 'y'), None)
        exp_col = next((c for c in df_test.columns if 'exp_out' in c), None)

        if not x_col or not y_col or not exp_col:
            print("  -> ERROR: Test table missing 'x', 'y' or 'exp_out' columns.")
            return

        # 预热 JIT
        dummy = jnp.array(0, dtype=self.dtype)
        self.run_kernel(dummy, dummy)

        for idx, row in df_test.iterrows():
            x_name = str(row[x_col]).strip()
            y_name = str(row[y_col]).strip()

            # 查找值
            val_x = value_registry.get(x_name)
            val_y = value_registry.get(y_name)

            if val_x is None or val_y is None:
                test_outs.append("ERR_UNK")
                results.append("Error: Unknown Input")
                continue

            try:
                # 执行运算
                res = self.run_kernel(val_x, val_y)
                res.block_until_ready()

                # 转换结果
                hex_res = jax_to_hex(res)
                test_outs.append(hex_res)

                # 对比
                exp_hex = str(row[exp_col]).strip()
                norm_act = hex_res.lower().replace('0x', '')
                norm_exp = exp_hex.lower().replace('0x', '')

                if norm_act == norm_exp:
                    results.append("PASS")
                else:
                    results.append(f"FAIL (Exp: {exp_hex})")

            except Exception as e:
                test_outs.append("ERR_EXEC")
                results.append(f"Exc: {str(e)}")

        df_test['test_out'] = test_outs
        df_test['result'] = results

        # 4. 保存为美化后的 Excel
        self.save_excel_report(output_file, df_def, df_test)

    def save_excel_report(self, out_path, df_def, df_test):
        try:
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                # 写入 Definition
                df_def.to_excel(writer, sheet_name='Result', index=False, startrow=0)

                # 写入 Test (留3行空隙)
                start_row_test = len(df_def) + 3
                df_test.to_excel(writer, sheet_name='Result', index=False, startrow=start_row_test)

                # --- 样式调整 ---
                workbook = writer.book
                sheet = writer.sheets['Result']

                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                red_font = Font(color="FF0000", bold=True)

                def set_border(min_r, max_r, min_c, max_c):
                    for row in sheet.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
                        for cell in row:
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='left')

                # 加边框
                set_border(1, len(df_def) + 1, 1, len(df_def.columns))
                test_table_start = start_row_test + 1
                test_table_end = test_table_start + len(df_test)
                set_border(test_table_start, test_table_end, 1, len(df_test.columns))

                # 调整列宽
                for col_idx in range(1, len(df_test.columns) + 1):
                    col_letter = get_column_letter(col_idx)
                    sheet.column_dimensions[col_letter].width = 30

                # 标红 FAIL
                result_col_idx = df_test.columns.get_loc('result') + 1
                for i in range(len(df_test)):
                    row_idx = test_table_start + 1 + i
                    cell = sheet.cell(row=row_idx, column=result_col_idx)
                    val_str = str(cell.value)

                    if val_str.startswith("FAIL") or "Error" in val_str:
                        for col in range(1, len(df_test.columns) + 1):
                            sheet.cell(row=row_idx, column=col).font = red_font

            # 打印统计
            pass_cnt = df_test['result'].value_counts().get('PASS', 0)
            total_cnt = len(df_test)
            print(f"  -> Saved to {out_path} (Pass: {pass_cnt}/{total_cnt})")

        except Exception as e:
            print(f"  -> Error saving Excel: {e}")


if __name__ == "__main__":
    # 使用示例

    # 1. 运行 FP32 测试
    tester_fp32 = TPUAddTester(mode='fp32')
    # 注意：这里直接读取包含 type/val 和 test cases 的完整 Excel 文件
    tester_fp32.run_test_file("fp32_add.xlsx", "fp32_add_result.xlsx")

    # 2. 运行 BF16 测试
    tester_bf16 = TPUAddTester(mode='bf16')
    tester_bf16.run_test_file("bf16_add.xlsx", "bf16_add_result.xlsx")