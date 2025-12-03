import os
import glob
import re
import pandas as pd
import numpy as np
import jax
import jax.numpy as jnp
from jax import lax
import logging

# os.environ['JAX_PLATFORM_NAME'] = 'tpu'

# 引入样式库
from openpyxl.styles import PatternFill, Font

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger()


# ==========================================
# 1. 通用工具：Hex 与 JAX 转换 (保持不变)
# ==========================================
def hex_to_jax(hex_str, dtype):
    """通用 Hex -> JAX Array 转换"""
    clean_hex = str(hex_str).strip().lower()
    if pd.isna(hex_str) or not clean_hex or clean_hex == 'nan':
        return jnp.array(0, dtype=dtype)
    if clean_hex.startswith('0x'): clean_hex = clean_hex[2:]
    clean_hex = clean_hex.replace(' ', '')  # Handle "FFFF FFFF"
    try:
        int_val = int(clean_hex, 16)
    except ValueError:
        return jnp.array(0, dtype=dtype)  # Fallback

    if dtype == jnp.float32:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint32), jnp.float32)
    elif dtype == jnp.bfloat16:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint16), jnp.bfloat16)
    elif dtype == jnp.int32:
        return jnp.array(int_val, dtype=jnp.int32)
    elif dtype == jnp.uint32:
        return jnp.array(int_val, dtype=jnp.uint32)
    elif dtype == jnp.int16:
        return jnp.array(int_val, dtype=jnp.int16)
    elif dtype == jnp.uint16:
        return jnp.array(int_val, dtype=jnp.uint16)

    return jnp.array(int_val, dtype=dtype)


def jax_to_hex(jax_val):
    """通用 JAX Array -> Hex 转换"""
    if jax_val.dtype == jnp.bool_:
        return "0x1" if jax_val else "0x0"

    if jax_val.dtype == jnp.float32:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint32)
        width = 8
    elif jax_val.dtype == jnp.bfloat16:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint16)
        width = 4
    elif jax_val.dtype in [jnp.int32, jnp.uint32]:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint32)
        width = 8
    elif jax_val.dtype in [jnp.int16, jnp.uint16]:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint16)
        width = 4
    elif jax_val.dtype == jnp.int8 or jax_val.dtype == jnp.uint8:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint8)
        width = 2
    else:
        return str(jax_val)

    try:
        int_val = int(val_uint)
    except:
        int_val = int(val_uint.item())
    return f"0x{int_val:0{width}X}"


# ==========================================
# 2. 算子注册表 (Op Registry) (保持不变)
# ==========================================
OP_REGISTRY = [
    (r".*_add.*", lambda x, y: x + y),
    (r".*_sub.*", lambda x, y: x - y),
    (r".*_mul.*", lambda x, y: x * y),
    (r".*_div.*", lambda x, y: x / y),
    (r".*_min.*", jnp.minimum),
    (r".*_max.*", jnp.maximum),
    (r".*_and\..*", lambda x, y: jnp.bitwise_and(x, y)),
    (r".*_or.*", lambda x, y: jnp.bitwise_or(x, y)),
    (r".*_xor.*", lambda x, y: jnp.bitwise_xor(x, y)),
    (r".*_eq.*", lambda x, y: x == y),
    (r".*_ne.*", lambda x, y: x != y),
    (r".*_gt.*", lambda x, y: x > y),
    (r".*_lt.*", lambda x, y: x < y),
    (r".*_ge.*", lambda x, y: x >= y),
    (r".*_le.*", lambda x, y: x <= y),
    (r".*f32_to_bf16.*", lambda x: lax.convert_element_type(x, jnp.bfloat16)),
    (r".*bf16_to_f32.*", lambda x: lax.convert_element_type(x, jnp.float32)),
    (r".*_ceil.*", jnp.ceil),
    (r".*_floor.*", jnp.floor),
    (r".*_trunc.*", jnp.trunc),
]


def get_op_from_filename(filename):
    fname = os.path.basename(filename).lower()
    for pattern, func in OP_REGISTRY:
        if re.match(pattern, fname):
            return func
    return None


def get_dtype_from_filename(filename):
    fname = os.path.basename(filename).lower()
    if 'f32' in fname: return jnp.float32
    if 'bf16' in fname: return jnp.bfloat16
    if 'u32' in fname: return jnp.uint32
    if 's32' in fname: return jnp.int32
    if 'u16' in fname: return jnp.uint16
    if 's16' in fname: return jnp.int16
    return jnp.float32


# ==========================================
# 3. 智能 CSV 加载器 (微调返回值)
# ==========================================
def parse_test_file(filepath):
    try:
        df_raw = pd.read_csv(filepath, header=None, engine='python')
    except Exception as e:
        logger.error(f"  [Error] Read failed: {e}")
        return None, None

    def_start_row = -1
    test_start_row = -1

    # 寻找分割行
    for i, row in df_raw.iterrows():
        row_str = " ".join([str(x).lower() for x in row.values])
        if 'type' in row_str and 'val' in row_str:
            def_start_row = i
        if ('test' in row_str or 'x' in row_str) and 'exp_out' in row_str:
            test_start_row = i

    if def_start_row == -1 or test_start_row == -1:
        logger.warning(f"  [Skip] Headers not found in {os.path.basename(filepath)}")
        return None, None

    # --- 提取 Definition 表 (自动剔除空列) ---
    header_row_def = df_raw.iloc[def_start_row]
    # 找到所有非 NaN 的列索引
    valid_cols_def = header_row_def.dropna().index
    # 只提取这些列
    df_def = df_raw.iloc[def_start_row + 1: test_start_row, valid_cols_def].copy()
    # 重命名列
    df_def.columns = header_row_def[valid_cols_def].astype(str).str.strip()
    df_def = df_def.dropna(how='all')  # 去掉纯空行

    # --- 提取 Test 表 (自动剔除空列) ---
    header_row_test = df_raw.iloc[test_start_row]
    valid_cols_test = header_row_test.dropna().index
    df_test = df_raw.iloc[test_start_row + 1:, valid_cols_test].copy()
    df_test.columns = header_row_test[valid_cols_test].astype(str).str.strip()
    df_test = df_test.dropna(how='all')

    return df_def, df_test


# ==========================================
# 4. 主执行逻辑 (核心修改)
# ==========================================
def run_single_file(filepath, output_dir):
    filename = os.path.basename(filepath)
    base_name = os.path.splitext(filename)[0]
    out_path = os.path.join(output_dir, f"RES_{base_name}.xlsx")

    logger.info(f"Processing {filename}...")

    op_func = get_op_from_filename(filename)
    if not op_func:
        logger.warning(f"  [Skip] No operator matched")
        return

    dtype = get_dtype_from_filename(filename)

    # 加载 (已经变紧凑了)
    df_def, df_test = parse_test_file(filepath)
    if df_def is None: return

    # --- 构建注册表 ---
    value_registry = {}
    # 模糊匹配列名
    type_col = next((c for c in df_def.columns if 'type' in c.lower()), None)
    val_col = next((c for c in df_def.columns if 'val' in c.lower()), None)

    if type_col and val_col:
        for _, row in df_def.iterrows():
            if pd.notna(row[type_col]) and pd.notna(row[val_col]):
                name = str(row[type_col]).strip()
                hex_val = str(row[val_col]).strip()
                value_registry[name] = hex_to_jax(hex_val, dtype)

    # --- 运行测试 ---
    jit_op = jax.jit(op_func)
    test_outs = []
    results = []

    x_col = next((c for c in df_test.columns if c.strip() == 'x'), None)
    y_col = next((c for c in df_test.columns if c.strip() == 'y'), None)
    exp_col = next((c for c in df_test.columns if 'exp_out' in c), None)

    if not x_col or not exp_col:
        logger.warning("  [Skip] Missing columns x or exp_out")
        return

    is_binary = (y_col is not None)

    for idx, row in df_test.iterrows():
        x_name = row[x_col]
        if pd.isna(x_name):
            test_outs.append("")
            results.append("")
            continue

        x_name = str(x_name).strip()
        val_x = value_registry.get(x_name)

        if val_x is None:
            test_outs.append("ERR_X")
            results.append("Error: Unknown X")
            continue

        try:
            if is_binary:
                y_name = str(row[y_col]).strip()
                val_y = value_registry.get(y_name)
                if val_y is None:
                    test_outs.append("ERR_Y")
                    results.append("Error: Unknown Y")
                    continue
                res = jit_op(val_x, val_y)
            else:
                res = jit_op(val_x)

            res.block_until_ready()
            hex_res = jax_to_hex(res)
            test_outs.append(hex_res)

            exp_hex = str(row[exp_col]).strip()
            norm_act = hex_res.lower().replace('0x', '')
            norm_exp = exp_hex.lower().replace('0x', '')

            if norm_act == norm_exp:
                results.append("PASS")
            else:
                results.append(f"FAIL (Exp: {exp_hex})")

        except Exception as e:
            test_outs.append("ERR")
            results.append(f"Exc: {str(e)}")

    df_test['test_out'] = test_outs
    df_test['result'] = results

    # --- 保存 Excel 并应用样式 ---
    try:
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            # 1. 写入 Definition (紧凑表格)
            df_def.to_excel(writer, sheet_name='Result', index=False, startrow=0)

            # 2. 写入 Test (紧凑表格)
            # 留 3 行空白
            start_row_test = len(df_def) + 3
            df_test.to_excel(writer, sheet_name='Result', index=False, startrow=start_row_test)

            # --- 样式调整 ---
            workbook = writer.book
            sheet = writer.sheets['Result']

            # 定义样式
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            red_font = Font(color="FF0000", bold=True)  # 纯红字，加粗

            # 辅助函数：给区域加边框
            def set_border(min_r, max_r, min_c, max_c):
                for row in sheet.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left')  # 左对齐比较好看

            # 给 Definition 表加边框
            set_border(1, len(df_def) + 1, 1, len(df_def.columns))

            # 给 Test 表加边框
            test_table_start = start_row_test + 1  # 1-based index including header
            test_table_end = test_table_start + len(df_test)
            set_border(test_table_start, test_table_end, 1, len(df_test.columns))

            # 标红逻辑 (Result 列)
            result_col_idx = df_test.columns.get_loc('result') + 1

            # 遍历 Test 表的数据行
            for i in range(len(df_test)):
                row_idx = test_table_start + 1 + i  # 跳过header
                cell = sheet.cell(row=row_idx, column=result_col_idx)
                val_str = str(cell.value)

                # 如果包含 FAIL 或 ERROR，只变字体颜色
                if val_str.startswith("FAIL") or "Error" in val_str:
                    # 将整行的字体变红
                    for col in range(1, len(df_test.columns) + 1):
                        sheet.cell(row=row_idx, column=col).font = red_font

        # 统计
        pass_cnt = results.count("PASS")
        total_cnt = len([r for r in results if r])
        logger.info(f"  -> Saved {out_path} (Pass: {pass_cnt}/{total_cnt})")

    except Exception as e:
        logger.error(f"  [Error] Save Excel failed: {e}")


def main():
    # --- TPU Check ---
    print("=" * 60)
    print("Checking Runtime Environment...")
    devices = jax.devices()
    print(f"Running on devices: {devices}")

    # 检查是否有 TPU 设备
    is_tpu = any('TPU' in str(d) for d in devices)
    if is_tpu:
        print("✅ SUCCESS: TPU detected. Running tests on TPU hardware.")
    else:
        print("⚠️ WARNING: Not running on TPU! Tests are running on CPU/GPU.")
        print("   Bitwise results (especially NaN/subnormals) may differ from hardware spec.")
    print("=" * 60 + "\n")

    input_dir = "./input_csvs"
    output_dir = "./test_results_xlsx"  # 使用新的输出目录区分

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    files = glob.glob(os.path.join(input_dir, "*.csv"))
    print(f"Found {len(files)} files.")

    for f in files:
        run_single_file(f, output_dir)


if __name__ == "__main__":
    main()