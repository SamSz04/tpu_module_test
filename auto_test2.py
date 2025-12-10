import os
import glob
import re
import pandas as pd
import numpy as np
import jax
import jax.numpy as jnp
from jax import lax
import logging
import ml_dtypes

# os.environ['JAX_PLATFORM_NAME'] = 'tpu'

# 引入样式库
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

jax.config.update("jax_enable_x64", True)
# 设置日志
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger()


def str_to_dtype(s):
    """将类型字符串转换为 JAX/ml_dtypes 类型"""
    s = s.lower()
    # FP8
    if 'f8e5' in s or 'e5m2' in s: return ml_dtypes.float8_e5m2
    if 'f8e4' in s or 'e4m3' in s: return ml_dtypes.float8_e4m3fn
    # FP4
    if 'f4e2' in s or 'e2m1' in s: return ml_dtypes.float4_e2m1fn
    # Standard Float
    if 'f32' in s: return jnp.float32
    if 'bf16' in s: return jnp.bfloat16
    if 'f16' in s or 'hf16' in s: return jnp.float16
    # Integer (Signed)
    if 's32' in s: return jnp.int32
    if 's16' in s: return jnp.int16
    if 's8' in s: return jnp.int8
    if 's4' in s: return ml_dtypes.int4
    if 's2' in s: return ml_dtypes.int2
    if 's1' in s: return ml_dtypes.int2 # fallback
    # Integer (Unsigned / Raw Bits)
    if 'u32' in s or 'b32' in s: return jnp.uint32
    if 'u16' in s or 'b16' in s: return jnp.uint16
    if 'u8' in s or 'b8' in s: return jnp.uint8
    if 'u4' in s or 'b4' in s: return ml_dtypes.uint4
    if 'u2' in s or 'b2' in s: return ml_dtypes.uint2
    if 'u1' in s or 'b1' in s: return ml_dtypes.uint2
    return None


def get_cvt_dtypes(filename):
    """解析文件名中的 Source 和 Target 类型"""
    fname = os.path.basename(filename).lower().replace('.csv', '')

    parts = []
    if '_to_' in fname:
        parts = fname.split('_to_')
    elif '->' in fname:
        parts = fname.split('->')
    elif fname.startswith('cvt_'):
        # 尝试移除 cvt_ 后，匹配第一个下划线
        content = fname[4:]
        # 暴力匹配：尝试所有已知类型作为前缀
        known_types = ['f32', 'bf16', 'f16', 'hf16', 's32', 'u32', 'b32',
                       's16', 'u16', 'b16', 's8', 'u8', 'b8', 's4', 'u4', 'b4',
                       's2', 'u2', 'b2', 'f8e5', 'e5m2', 'f8e4', 'e4m3', 'f4e2', 'e2m1']
        # 按长度降序排，防止前缀冲突 (如 f8 匹配了 f8e5)
        known_types.sort(key=len, reverse=True)

        for t in known_types:
            if content.startswith(t + '_'):
                src = t
                dst = content[len(t) + 1:]
                # 去掉可能的后缀 (如 _rs)
                if '_rs' in dst: dst = dst.replace('_rs', '')
                parts = [src, dst]
                break

    if len(parts) == 2:
        return str_to_dtype(parts[0]), str_to_dtype(parts[1])
    return None, None


def get_default_dtype(filename):
    """旧的单类型推断 (用于非 CVT 文件)"""
    # 复用 str_to_dtype 的逻辑，简单匹配文件名包含的字符串
    fname = os.path.basename(filename).lower()

    # bf16_clampgez 和 bf16_clamps 是 32-bit Packed 格式
    if 'bf16' in fname and ('clampgez' in fname or 'clamps' in fname):
        return jnp.uint32

    known_types = ['f32', 'bf16', 'f16', 's32', 'u32', 's16', 'u16', 's8', 'u8', 's4', 'u4', 'f8e5', 'e5m2', 'f8e4',
                   'e4m3', 'f4e2', 'e2m1']
    known_types.sort(key=len, reverse=True)
    for t in known_types:
        if t in fname: return str_to_dtype(t)
    return jnp.float32


# ==========================================
# 0. Helper Kernels
# ==========================================
def op_int_carry(x, y):
    """计算整数加法的进位 (Carry Out)"""
    # 提升到更高精度计算，看高位是否有值
    # 假设输入是 u32/u16，提升到 u64
    x_64 = x.astype(jnp.uint64)
    y_64 = y.astype(jnp.uint64)
    res_64 = x_64 + y_64
    # 检查是否溢出 (对于 u32，最大值是 2^32-1)
    limit = jnp.iinfo(x.dtype).max
    return (res_64 > limit).astype(x.dtype) # 返回 0 或 1

def _to_ord_bits(x):
    """
    将浮点数 x 转换为一个保序的 unsigned integer。
    映射规则:
    - 如果 x >= 0 (MSB=0): x ^ 0x8000... (翻转符号位，移到高半区)
    - 如果 x < 0  (MSB=1): ~x (按位取反，反转顺序并移到低半区)
    """
    dtype = x.dtype

    # 1. Bitcast 为无符号整数
    if dtype == jnp.float32:
        u = lax.bitcast_convert_type(x, jnp.uint32)
        # MSB 掩码 (0x80000000)
        sign_bit = jnp.array(0x80000000, dtype=jnp.uint32)
        mask_neg = (u >> 31) == 1
    elif dtype == jnp.bfloat16:
        u = lax.bitcast_convert_type(x, jnp.uint16)
        # MSB 掩码 (0x8000)
        sign_bit = jnp.array(0x8000, dtype=jnp.uint16)
        mask_neg = (u >> 15) == 1
    elif hasattr(jnp, 'float16') and dtype == jnp.float16:  # 兼容 fp16
        u = lax.bitcast_convert_type(x, jnp.uint16)
        sign_bit = jnp.array(0x8000, dtype=jnp.uint16)
        mask_neg = (u >> 15) == 1
    else:
        # 非浮点数直接返回 (假设已经是整数)
        return x

    # 2. 根据符号位应用映射
    # 如果是负数：取反 (~u)
    # 如果是正数：翻转符号位 (u ^ sign_bit)
    ord_u = jnp.where(mask_neg, jnp.bitwise_not(u), jnp.bitwise_xor(u, sign_bit))

    return ord_u


def op_total_order_lt(x, y):
    """Total Order Less Than (<)"""
    return _to_ord_bits(x) < _to_ord_bits(y)


def op_total_order_le(x, y):
    """Total Order Less Equal (<=)"""
    return _to_ord_bits(x) <= _to_ord_bits(y)

def op_clz(x):
    """Count Leading Zeros"""
    return lax.clz(x)

def op_pcnt(x):
    """Population Count (统计 1 的个数)"""
    return lax.population_count(x)


def op_remap(x, y):
    """
    硬件 Remap 指令完全实现版 (修正版)。
    """
    dtype = x.dtype

    # ==========================================
    # 1. 提取 x 特征 & 计算 Bit Position
    # ==========================================
    if dtype == jnp.float32:
        u = lax.bitcast_convert_type(x, jnp.uint32)
        sign = u >> 31
        exp = (u >> 23) & 0xFF
        mant = u & 0x7FFFFFFF

        MASK_INF_EXP = 0xFF

        # --- 常量定义 (Hex) ---
        HEX_POS_ZERO = 0x00000000
        HEX_NEG_ZERO = 0x80000000
        HEX_POS_ONE = 0x3F800000  # 1.0
        HEX_NEG_ONE = 0xBF800000  # -1.0
        HEX_POS_INF = 0x7F800000
        HEX_NEG_INF = 0xFF800000
        HEX_QNAN = 0x7FC00000

        # 9: FLP_MIN -> Max Negative Finite (-Max)
        HEX_MIN_FLP = 0xFF7FFFFF
        # 10: FLP_MAX -> Max Positive Finite
        HEX_MAX_FLP = 0x7F7FFFFF

        HEX_MAX_INT = 0xFFFFFFFF
        HEX_MIN_INT = 0x80000000
        HEX_POS_ONE_I = 0x00000001  # 整数 1
        HEX_NEG_ONE_I = 0xFFFFFFFF  # 整数 -1

    else:  # bf16
        u = lax.bitcast_convert_type(x, jnp.uint16)
        sign = u >> 15
        exp = (u >> 7) & 0xFF
        mant = u & 0x7F

        MASK_INF_EXP = 0xFF

        HEX_POS_ZERO = 0x0000
        HEX_NEG_ZERO = 0x8000
        HEX_POS_ONE = 0x3F80
        HEX_NEG_ONE = 0xBF80
        HEX_POS_INF = 0x7F80
        HEX_NEG_INF = 0xFF80
        HEX_QNAN = 0x7FC0

        # 9: FLP_MIN -> Max Negative Finite
        HEX_MIN_FLP = 0xFF7F
        # 10: FLP_MAX -> Max Positive Finite
        HEX_MAX_FLP = 0x7F7F

        HEX_MAX_INT = 0xFFFF
        HEX_MIN_INT = 0x8000
        HEX_POS_ONE_I = 0x0001
        HEX_NEG_ONE_I = 0xFFFF


    # 逻辑判断
    is_zero = (exp == 0) & (mant == 0)
    is_sub = (exp == 0) & (mant != 0)
    is_inf = (exp == MASK_INF_EXP) & (mant == 0)
    is_nan = (exp == MASK_INF_EXP) & (mant != 0)
    is_norm = (exp != 0) & (exp != MASK_INF_EXP)

    # 查表 1: 确定 y 的移位量 (Bit Position)
    # 规则: -Inf->0, -Norm->4, -Sub/-0->8, +Sub/+0->12, +Norm->16, +Inf->20, NaN->24
    bitpos = jnp.select([
        is_nan,  # 24
        (sign == 0) & is_inf,  # 20
        (sign == 0) & is_norm,  # 16
        (sign == 0) & (is_sub | is_zero),  # 12
        (sign == 1) & (is_sub | is_zero),  # 8
        (sign == 1) & is_norm,  # 4
        (sign == 1) & is_inf,  # 0
    ], [24, 20, 16, 12, 8, 4, 0], default=0)

    # ==========================================
    # 2. 获取 Opcode
    # ==========================================
    y_int = lax.bitcast_convert_type(y, jnp.uint32) if y.dtype == jnp.float32 else y.astype(jnp.uint32)
    opcode = jnp.right_shift(y_int, bitpos.astype(jnp.uint32)) & 0xF

    # ==========================================
    # 3. 构造常量池
    # ==========================================
    def mk_const(hex_val, target_dtype):
        if target_dtype == jnp.float32:
            return lax.bitcast_convert_type(jnp.array(hex_val, dtype=jnp.uint32), jnp.float32)
        else:
            return lax.bitcast_convert_type(jnp.array(hex_val, dtype=jnp.uint16), jnp.bfloat16)

    c_pos_zero = mk_const(HEX_POS_ZERO, dtype)
    c_neg_zero = mk_const(HEX_NEG_ZERO, dtype)
    c_pos_one = mk_const(HEX_POS_ONE, dtype)
    c_neg_one = mk_const(HEX_NEG_ONE, dtype)
    c_pos_inf = mk_const(HEX_POS_INF, dtype)
    c_neg_inf = mk_const(HEX_NEG_INF, dtype)
    c_qnan = mk_const(HEX_QNAN, dtype)

    # 修正后的 9 和 10
    c_min_flp = mk_const(HEX_MIN_FLP, dtype)  # FF7FFFFF
    c_max_flp = mk_const(HEX_MAX_FLP, dtype)  # 7F7FFFFF

    # 11 和 12 (INT_MAX/MIN):
    c_int_max = mk_const(HEX_MAX_INT, dtype)
    c_int_min = mk_const(HEX_MIN_INT, dtype)

    c_pos_one_i = mk_const(HEX_POS_ONE_I, dtype)
    c_neg_one_i = mk_const(HEX_NEG_ONE_I, dtype)

    # ==========================================
    # 4. 执行 Opcode 映射 (0~15)
    # ==========================================
    res = jnp.select([
        opcode == 0,  # x
        opcode == 1,  # +zero
        opcode == 2,  # -zero
        opcode == 3,  # +1
        opcode == 4,  # -1
        opcode == 5,  # -inf
        opcode == 6,  # +inf
        opcode == 7,  # -x
        opcode == 8,  # qNan
        opcode == 9,  # FLP_MIN ->  HEX_MIN_FLP
        opcode == 10,  # FLP_MAX -> HEX_MAX_FLP
        opcode == 11,  # INT_MAX
        opcode == 12,  # INT_MIN
        opcode == 13,  # abs(x)
        opcode == 14,  # +1 (Duplicate of 3)
        opcode == 15  # -1 (Duplicate of 4)
    ], [
        x,  # 0
        c_pos_zero,  # 1
        c_neg_zero,  # 2
        c_pos_one,  # 3
        c_neg_one,  # 4
        c_neg_inf,  # 5
        c_pos_inf,  # 6
        -x,  # 7
        c_qnan,  # 8
        c_min_flp,  # 9
        c_max_flp,  # 10
        c_int_max,  # 11
        c_int_min,  # 12
        jnp.abs(x),  # 13
        c_pos_one_i,  # 14
        c_neg_one_i  # 15
    ], default=x)

    return res


def op_class(x, y):
    """
    硬件 Class 指令模拟。
    1. 计算 x 的类别 ID (0~9)。
    2. 返回 (y >> ID) & 1。
    """
    # 1. 提取符号、指数、尾数
    if x.dtype == jnp.float32:
        u = lax.bitcast_convert_type(x, jnp.uint32)
        sign = u >> 31
        exp = (u >> 23) & 0xFF
        mant = u & 0x7FFFFFFF
    else:  # bf16 / f16
        u = lax.bitcast_convert_type(x, jnp.uint16)
        sign = u >> 15
        exp = (u >> 7) & 0xFF
        mant = u & 0x7F

    is_zero = (exp == 0) & (mant == 0)
    is_sub = (exp == 0) & (mant != 0)
    is_inf = (exp == 0xFF) & (mant == 0)
    is_nan = (exp == 0xFF) & (mant != 0)
    is_norm = (exp != 0) & (exp != 0xFF)

    # 2. 映射到 Class ID (0-9)
    # 注意: JAX select 是从后往前匹配优先级，我们要确保互斥或者顺序正确
    class_id = jnp.select([
        (sign == 1) & is_nan,  # 0: -NaN
        (sign == 1) & is_inf,  # 1: -Inf
        (sign == 1) & is_norm,  # 2: -Normal
        (sign == 1) & is_sub,  # 3: -Sub
        (sign == 1) & is_zero,  # 4: -Zero
        (sign == 0) & is_zero,  # 5: +Zero
        (sign == 0) & is_sub,  # 6: +Sub
        (sign == 0) & is_norm,  # 7: +Normal
        (sign == 0) & is_inf,  # 8: +Inf
        (sign == 0) & is_nan  # 9: +NaN
    ], [0, 1, 2, 3, 4, 5, 6, 7, 8, 9], default=0)

    # 3. 查表 y
    # y 可能是 float 类型传入的，需转为 integer
    y_int = lax.bitcast_convert_type(y, jnp.uint32) if y.dtype == jnp.float32 else y.astype(jnp.uint32)

    # Shift and Mask
    # result = (y >> class_id) & 1
    # 注意: class_id 是 array，JAX 支持 vector shift
    return (jnp.right_shift(y_int, class_id.astype(jnp.uint32)) & 1).astype(jnp.bool_)


# --- Clamp 系列 ---

def op_clampgez(x, y):
    """
    Clamp Greater Equal Zero (ReLU variant)
    Rule: max(min(x, y), +0)
    """
    # 确保 0 是同类型的 +0
    zero = jnp.array(0.0, dtype=x.dtype)
    return jnp.maximum(jnp.minimum(x, y), zero)


def op_clamps(x, y):
    """
    Clamp Symmetric (Saturate)
    Rule: min(max(x, -y), max(y, +0))
    """
    zero = jnp.array(0.0, dtype=x.dtype)
    # 限制在 [-y, max(y, 0)] 之间
    return jnp.minimum(jnp.maximum(x, -y), jnp.maximum(y, zero))


def op_packed_bf16_clampgez(x, y):
    """
    BF16 Packed ClampGEZ (SIMD)
    Input: x, y as uint32 (each contains 2 bf16s)
    Logic:
      High: max(min(x_h, y_h), 0)
      Low:  max(min(x_l, y_l), 0)
    """
    # 1. Unpack uint32 -> 2x bf16
    x_h = lax.bitcast_convert_type((x >> 16).astype(jnp.uint16), jnp.bfloat16)
    x_l = lax.bitcast_convert_type((x & 0xFFFF).astype(jnp.uint16), jnp.bfloat16)
    y_h = lax.bitcast_convert_type((y >> 16).astype(jnp.uint16), jnp.bfloat16)
    y_l = lax.bitcast_convert_type((y & 0xFFFF).astype(jnp.uint16), jnp.bfloat16)

    # 2. Compute
    zero = jnp.array(0.0, dtype=jnp.bfloat16)
    res_h = jnp.maximum(jnp.minimum(x_h, y_h), zero)
    res_l = jnp.maximum(jnp.minimum(x_l, y_l), zero)

    # 3. Pack bf16 -> uint32
    r_h = lax.bitcast_convert_type(res_h, jnp.uint16).astype(jnp.uint32)
    r_l = lax.bitcast_convert_type(res_l, jnp.uint16).astype(jnp.uint32)
    return (r_h << 16) | r_l


def op_packed_bf16_clamps(x, y):
    """
    BF16 Packed Clamps (SIMD)
    Input: x, y as uint32
    Logic:
      High: min(max(x_h, -y_h), max(y_h, 0))
      Low:  min(max(x_l, -y_l), max(y_l, 0))
    """
    # 1. Unpack
    x_h = lax.bitcast_convert_type((x >> 16).astype(jnp.uint16), jnp.bfloat16)
    x_l = lax.bitcast_convert_type((x & 0xFFFF).astype(jnp.uint16), jnp.bfloat16)
    y_h = lax.bitcast_convert_type((y >> 16).astype(jnp.uint16), jnp.bfloat16)
    y_l = lax.bitcast_convert_type((y & 0xFFFF).astype(jnp.uint16), jnp.bfloat16)

    # 2. Compute
    zero = jnp.array(0.0, dtype=jnp.bfloat16)

    # Upper bound needs to be max(y, 0) according to spec
    limit_h = jnp.maximum(y_h, zero)
    limit_l = jnp.maximum(y_l, zero)

    res_h = jnp.minimum(jnp.maximum(x_h, -y_h), limit_h)
    res_l = jnp.minimum(jnp.maximum(x_l, -y_l), limit_l)

    # 3. Pack
    r_h = lax.bitcast_convert_type(res_h, jnp.uint16).astype(jnp.uint32)
    r_l = lax.bitcast_convert_type(res_l, jnp.uint16).astype(jnp.uint32)
    return (r_h << 16) | r_l


def op_clampa_f32(x, y_h_bits, y_l_bits):
    """
    F32 Asymmetric Clamp
    Input: x(f32), y_h(bf16 bits), y_l(bf16 bits)
    Rule: max(min(x, y_h_f32), y_l_f32)
    Note: y_h/y_l 在 CSV 中是 bitwise value，需要左移 16 位转为 f32
    """
    # 将 bf16 bits 扩展为 f32 (padding 0 at LSB)
    y_h = lax.bitcast_convert_type(y_h_bits.astype(jnp.uint32) << 16, jnp.float32)
    y_l = lax.bitcast_convert_type(y_l_bits.astype(jnp.uint32) << 16, jnp.float32)

    return jnp.maximum(jnp.minimum(x, y_h), y_l)


def op_clampa_bf16(x_h, x_l, y_h, y_l):
    """
    BF16 Packed Asymmetric Clamp
    Input: x={x_h, x_l}, y={y_h, y_l} (Global Bounds)
    Rule: Clamp both x_h and x_l using range [y_l, y_h]
    """
    out_h = jnp.maximum(jnp.minimum(x_h, y_h), y_l)
    out_l = jnp.maximum(jnp.minimum(x_l, y_h), y_l)
    return out_h, out_l


def op_weird(x, y=None):
    """
    Weird Operation: Detects Inf or NaN.
    Output 1 if x is Inf or NaN, else 0.
    """
    # 1. Check for Infinity (positive or negative)
    is_inf = jnp.isinf(x)

    # 2. Check for NaN
    is_nan = jnp.isnan(x)

    # 3. Combine conditions (Logical OR)
    res = is_inf | is_nan

    # 4. Return as bool so jax_to_hex outputs "0x1" or "0x0"
    return res.astype(jnp.bool_)


# ==========================================
# 1. 通用工具：Hex 与 JAX 转换
# ==========================================
def hex_to_jax(hex_str, dtype):
    """通用 Hex -> JAX Array 转换"""
    clean_hex = str(hex_str).strip().lower()
    if pd.isna(hex_str) or not clean_hex or clean_hex == 'nan':
        # 空值默认处理：对于复杂类型，先创建 uint8 0，再 bitcast
        if dtype in [ml_dtypes.float4_e2m1fn, ml_dtypes.int4, ml_dtypes.uint4]:
            return jnp.zeros((), dtype=dtype)
        return jnp.array(0, dtype=dtype)

    if clean_hex.startswith('0x'): clean_hex = clean_hex[2:]
    clean_hex = clean_hex.replace(' ', '')
    try:
        int_val = int(clean_hex, 16)
    except ValueError:
        return jnp.array(0, dtype=dtype)

    # ============================================
    # ml_dtypes 特殊处理 (FP8, FP4, int4)
    # ============================================

    # 1. FP8 类型 (E5M2, E4M3)
    if dtype in [ml_dtypes.float8_e5m2, ml_dtypes.float8_e4m3fn]:
        # 先转为 uint8 容器，再 bitcast 为 fp8
        uint8_val = jnp.array(int_val, dtype=jnp.uint8)
        return lax.bitcast_convert_type(uint8_val, dtype)

    # 2. FP4 类型 (f4e2 -> float4_e2m1fn)
    # 这一步比较 tricky，JAX 的 bitcast 可能还不完全支持 <8bit 的类型互转
    # 我们可以利用 numpy 的 view 来构造初始值
    if dtype == ml_dtypes.float4_e2m1fn:
        # 在 Host 端用 numpy 构造 bit pattern
        # 注意：int_val 是 0~15 的整数，代表 raw bits
        np_val = np.array(int_val, dtype=np.uint8).view(ml_dtypes.float4_e2m1fn)
        return jnp.array(np_val)

    # int4 (Signed): 需要处理补码 (e.g. 0xF -> -1)
    if dtype == ml_dtypes.int4:
        # 手动补码转换: if >= 8 (0b1000), subtract 16
        if int_val >= 8: int_val -= 16
        return jnp.array(int_val, dtype=dtype)

    if dtype == ml_dtypes.uint4:
        return jnp.array(int_val, dtype=dtype)

    # ============================================
    # 标准 JAX 类型
    # ============================================
    if dtype == jnp.float32:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint32), jnp.float32)
    elif dtype == jnp.bfloat16:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint16), jnp.bfloat16)
    elif hasattr(jnp, 'float16') and dtype == jnp.float16:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint16), jnp.float16)

    # Signed Integers: 必须处理补码！
    # 策略：先创建 unsigned array，然后 bitcast/view 为 signed
    if dtype == jnp.int32:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint32), jnp.int32)
    elif dtype == jnp.int16:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint16), jnp.int16)
    elif dtype == jnp.int8:
        return lax.bitcast_convert_type(jnp.array(int_val, dtype=jnp.uint8), jnp.int8)

    # 标准整数
    return jnp.array(int_val, dtype=dtype)


def jax_to_hex(jax_val):
    dtype = jax_val.dtype
    if dtype == jnp.bool_: return "0x1" if jax_val else "0x0"

    width = 8
    val_uint = jax_val  # Default

    # --- ml_dtypes 处理 ---
    if dtype in [ml_dtypes.float8_e5m2, ml_dtypes.float8_e4m3fn]:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint8)
        width = 2
    elif dtype == ml_dtypes.float4_e2m1fn:
        # JAX 可能无法直接 bitcast float4 -> uint8
        # 转回 numpy 处理 view
        val_np = np.array(jax_val).view(np.uint8)
        return f"0x{int(val_np):01X}"  # 1 hex char
    elif dtype in [ml_dtypes.int4, ml_dtypes.uint4]:
        width = 1
        val_uint = jax_val

    # --- 标准类型 ---
    elif dtype == jnp.float32:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint32)
        width = 8
    elif dtype == jnp.bfloat16:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint16)
        width = 4
    elif hasattr(jnp, 'float16') and dtype == jnp.float16:
        val_uint = lax.bitcast_convert_type(jax_val, jnp.uint16)
        width = 4
    elif dtype in [jnp.int32, jnp.uint32]:
        width = 8
    elif dtype in [jnp.int16, jnp.uint16]:
        width = 4
    elif dtype in [jnp.int8, jnp.uint8]:
        width = 2

    try:
        # int_val = int(val_uint)
        int_val = int(np.array(val_uint).item())

        # 掩码处理负数 (例如 int4 的 -1 是 0xF)
        if int_val < 0:
            mask = (1 << (width * 4)) - 1
            int_val = int_val & mask
    except Exception as e:
        return f"ERR_HEX: {str(e)}"

    return f"0x{int_val:0{width}X}"


# ==========================================
# 2. 算子注册表 (Op Registry)
# ==========================================
OP_REGISTRY = [
    # --------------------------
    # 1. 基础运算 (Binary)
    # --------------------------
    (r".*_add.*", jnp.add),
    (r".*_sub.*", jnp.subtract),
    (r".*_mul.*", jnp.multiply),
    # (r".*_div.*", lambda x, y: x * y),
    (r".*_min.*", jnp.minimum),
    (r".*_max.*", jnp.maximum),

    # --------------------------
    # 2. 位运算 & 移位
    # --------------------------
    (r".*_and\..*", jnp.bitwise_and),
    (r".*_andn.*", lambda x, y: jnp.bitwise_and(jnp.bitwise_not(x), y)),  # u32_andn = ~x & y
    (r".*_or.*", jnp.bitwise_or),
    (r".*_xor.*", jnp.bitwise_xor),

    (r".*_clz.*", op_clz),  # Count Leading Zeros
    (r".*_pcnt.*", op_pcnt),  # Population Count
    (r".*_nez.*", lambda x: (x != 0).astype(x.dtype)),  # Not Equal Zero

    # RNE (Round Nearest Even) 是默认模式
    (r".*_rtne.*", lambda x: lax.round(x, lax.RoundingMethod.TO_NEAREST_EVEN)),
    # RTNA (Round Nearest Away)
    (r".*_rtna.*", lambda x: lax.round(x, lax.RoundingMethod.AWAY_FROM_ZERO)),

    # 移位 (注意：JAX 的 shift 行为取决于输入是 signed 还是 unsigned)
    # SHLL/SHLA (左移通常一样)
    (r".*_shll.*", jnp.left_shift),
    (r".*_shla.*", jnp.left_shift),
    # SHRL (逻辑右移)
    # 对于unsigned， right_shift 是逻辑右移，对于 signed，是算术右移
    (r".*_shrl.*", lax.shift_right_logical),
    # SHRA (算术右移)
    (r".*_shra.*", lax.shift_right_arithmetic),

    # --------------------------
    # 3. 比较运算 (Compare)
    # --------------------------
    (r".*_eq.*", jnp.equal),
    (r".*_ne.*", jnp.not_equal),
    (r".*_gt.*", jnp.greater),
    (r".*_lt.*", jnp.less),
    (r".*_ge.*", jnp.greater_equal),
    (r".*_le.*", jnp.less_equal),

    # Total Order (全序比较) - 硬件可能有特殊指令
    (r".*_le_to.*", op_total_order_le),
    (r".*_lt_to.*", op_total_order_lt),

    # 进位检测
    (r".*int_carry.*", op_int_carry),

    # 取整
    (r".*_ceil.*",  jnp.ceil),
    (r".*_floor.*", jnp.floor),
    (r".*_trunc.*", jnp.trunc),

    (r".*_remap.*", op_remap),
    (r".*_class.*", op_class),

    # 特殊 Packed BF16 算子
    (r".*bf16_clampgez.*", op_packed_bf16_clampgez),
    (r".*bf16_clamps.*",   op_packed_bf16_clamps),

    (r".*clampgez.*", op_clampgez),
    (r".*clamps.*", op_clamps),
    # Clampa 比较特殊，在 Main Loop 里单独处理，这里也可以放个占位
    (r".*clampa.*", lambda x: x),

    (r".*_weird_1.*", op_weird),
    (r".*_weird_2.*", op_weird),
    # --------------------------
    # 4. 类型转换 (Convert / Cast)
    # --------------------------
    # --- 浮点互转 (Float <-> Float) ---
    # --- 浮点互转 ---
    (r".*f32_to_bf16.*", lambda x: x.astype(jnp.bfloat16)),
    (r".*bf16_to_f32.*", lambda x: x.astype(jnp.float32)),
    (r".*f32_to_f16.*", lambda x: x.astype(jnp.float16)),
    (r".*cvt_f16_f32.*", lambda x: x.astype(jnp.float32)),
    (r".*cvt_bf16_f32.*", lambda x: x.astype(jnp.float32)),

    # --- FP4 (f4e2) 转换 ---
    # JAX + ml_dtypes 现在支持 FP4 转换了！
    (r".*cvt_bf16_f4e2.*", lambda x: x.astype(ml_dtypes.float4_e2m1fn)),
    (r".*f4e2_to_bf16.*", lambda x: x.astype(jnp.bfloat16)),  # 反向转换
    (r".*f4e2_to_f32.*", lambda x: x.astype(jnp.float32)),

    # --- FP8 转换 ---
    (r".*to_e5m2.*", lambda x: x.astype(ml_dtypes.float8_e5m2)),
    (r".*to_e4m3.*", lambda x: x.astype(ml_dtypes.float8_e4m3fn)),

    # --- 整数位数转换 (Signed) ---
    (r".*cvt_s16_s32.*", lambda x: x.astype(jnp.int32)),
    (r".*cvt_s8_s32.*", lambda x: x.astype(jnp.int32)),
    # Native int4 support!
    (r".*cvt_s4_s8.*", lambda x: x.astype(jnp.int8)),  # int4 -> int8 (Sign Ext)
    (r".*cvt_s2_s4.*", lambda x: x.astype(ml_dtypes.int4)),  # int2 -> int4 (Sign Ext)
    (r".*cvt_s1_s2.*", lambda x: x.astype(ml_dtypes.int2)),  # S1 -> int2

    # --- 位宽转换 (Raw Bits / Unsigned) ---
    (r".*cvt_b32_b16.*", lambda x: x.astype(jnp.uint16)),
    (r".*cvt_b16_b8.*", lambda x: x.astype(jnp.uint8)),
    # B8 -> B4 (使用 native uint4)
    (r".*cvt_b8_b4.*", lambda x: x.astype(ml_dtypes.uint4)),
    # B4 -> B8 (Zero Ext)
    (r".*cvt_b4_b8.*", lambda x: x.astype(jnp.uint8)),
    # B4 -> B2 (Truncation)
    (r".*cvt_b4_b2.*", lambda x: x.astype(ml_dtypes.uint2)),
    # B2 -> B4 (Zero Ext)
    (r".*cvt_b2_b4.*", lambda x: x.astype(ml_dtypes.uint4)),
    # B2 -> B1 (使用 uint2 模拟 B1，或截断)
    (r".*cvt_b2_b1.*", lambda x: x.astype(ml_dtypes.uint2)),  # 暂时 map 到 u2

    # --- Float <-> Int/u4/s4 ---
    (r".*f32_to_s32.*", lambda x: x.astype(jnp.int32)),
    (r".*f32_to_s4.*", lambda x: x.astype(ml_dtypes.int4)),  # 真正的 float->int4
    (r".*s32_to_f32.*", lambda x: x.astype(jnp.float32)),
    (r".*to_u8.*", lambda x: x.astype(jnp.uint8)),
    (r".*to_s8.*", lambda x: x.astype(jnp.int8)),
    (r".*to_u4.*", lambda x: x.astype(ml_dtypes.uint4)),
    (r".*to_s4.*", lambda x: x.astype(ml_dtypes.int4)),
    (r".*to_bf16.*", lambda x: x.astype(jnp.bfloat16)),
]


def get_op_from_filename(filename):
    fname = os.path.basename(filename).lower()
    for pattern, func in OP_REGISTRY:
        if re.match(pattern, fname):
            return func
    return None


# ==========================================
# 3. 智能 CSV 加载器 (微调返回值)
# ==========================================
def parse_test_file(filepath):
    try:
        df_raw = pd.read_csv(filepath, header=None, engine='python')
    except Exception as e:
        logger.error(f"  [Error] Read failed: {e}")
        return None, None

    def_start_row = -1
    test_start_row = -1

    for i, row in df_raw.iterrows():
        row_str = " ".join([str(x).lower() for x in row.values])
        if 'type' in row_str and 'val' in row_str:
            def_start_row = i
        if ('test' in row_str or 'x' in row_str) and 'exp_out' in row_str:
            test_start_row = i

    if def_start_row == -1 or test_start_row == -1:
        logger.warning(f"  [Skip] Headers not found in {os.path.basename(filepath)}")
        return None, None

    # --- 提取 Definition ---
    header_row_def = df_raw.iloc[def_start_row]
    valid_cols_def = header_row_def.dropna().index

    # 截取两个 header 之间的内容
    df_def = df_raw.iloc[def_start_row + 1: test_start_row, valid_cols_def].copy()
    df_def.columns = header_row_def[valid_cols_def].astype(str).str.strip()

    # 关键修复 2: 过滤掉内容为 "test" 的杂行
    # 只要第一列包含 "test" 字样，就认为是分隔行，扔掉
    if not df_def.empty:
        df_def = df_def.dropna(how='all')

        first_col = df_def.columns[0]
        mask_junk = df_def[first_col].astype(str).str.strip().str.lower() == 'test'
        df_def = df_def[~mask_junk]


    # --- 提取 Test ---
    header_row_test = df_raw.iloc[test_start_row]
    valid_cols_test = header_row_test.dropna().index
    df_test = df_raw.iloc[test_start_row + 1:, valid_cols_test].copy()
    df_test.columns = header_row_test[valid_cols_test].astype(str).str.strip()
    df_test = df_test.dropna(how='all')

    return df_def, df_test


# ==========================================
# 4. 主执行逻辑 (核心修改)
# ==========================================
def run_single_file(filepath, output_dir):
    filename = os.path.basename(filepath)
    base_name = os.path.splitext(filename)[0]
    out_path = os.path.join(output_dir, f"RES_{base_name}.xlsx")

    logger.info(f"Processing {filename}...")

    # 1. 双类型推断
    src_dtype, dst_dtype = get_cvt_dtypes(filename)

    if src_dtype and dst_dtype:
        dtype = src_dtype  # 加载 Definition 表用 Source 类型
        is_cvt = True
    else:
        dtype = get_default_dtype(filename)
        is_cvt = False

    # --- 2. 解析 & 注册 ---
    df_def, df_test = parse_test_file(filepath)
    if df_def is None: return

    # 2. 构建注册表
    value_registry = {}
    type_col = next((c for c in df_def.columns if 'type' in c.lower()), None)
    val_col = next((c for c in df_def.columns if 'val' in c.lower()), None)
    if type_col and val_col:
        for _, row in df_def.iterrows():
            if pd.notna(row[type_col]) and pd.notna(row[val_col]):
                name = str(row[type_col]).strip()
                hex_val = str(row[val_col]).strip()
                # 注意：clampa 的 y_high/low 也是用这个 dtype 读取的，如果是 f32 clampa, y 是 bf16 bits
                # 这里统一按 filename dtype 读取，后续 kernel 里再 bitcast
                value_registry[name] = hex_to_jax(hex_val, dtype)

    # 3. 准备运行
    op_func = get_op_from_filename(filename)
    test_outs = []
    results = []

    if is_cvt:
        def dynamic_cvt(x, y=None):
            # Try/Except 增强鲁棒性
            try:
                return x.astype(dst_dtype)
            except:
                return lax.convert_element_type(x, dst_dtype)

        op_func = dynamic_cvt

    if op_func is None:
        logger.warning("  [Skip] No operator matched")
        return

    # 辅助函数：安全查表
    def get_val(name):
        return value_registry.get(str(name).strip())

    # ==========================================
    # 分支 A: 处理 Clampa (特殊多列结构)
    # ==========================================
    if 'clampa' in filename:
        # Detect F32 or BF16 mode based on columns
        is_bf16_packed = 'x_high' in df_test.columns
        exp_col = next((c for c in df_test.columns if 'exp_out' in c), None)

        for idx, row in df_test.iterrows():
            try:
                if is_bf16_packed:
                    # BF16 Mode: x_h, x_l, y_h, y_l
                    vh_x = get_val(row['x_high'])
                    vl_x = get_val(row['x_low'])
                    vh_y = get_val(row['y_high'])
                    vl_y = get_val(row['y_low'])

                    if any(v is None for v in [vh_x, vl_x, vh_y, vl_y]):
                        raise ValueError("Unknown Value")

                    # 运行 Kernel
                    res_h, res_l = op_clampa_bf16(vh_x, vl_x, vh_y, vl_y)

                    # Pack 结果: (h << 16) | l
                    u_h = lax.bitcast_convert_type(res_h, jnp.uint16).astype(jnp.uint32)
                    u_l = lax.bitcast_convert_type(res_l, jnp.uint16).astype(jnp.uint32)
                    res_packed = (u_h << 16) | u_l
                    hex_res = f"0x{int(res_packed):08X}"

                else:
                    # F32 Mode: x, y_high, y_low
                    vx = get_val(row['x'])
                    vy_h = get_val(row['y_high'])
                    vy_l = get_val(row['y_low'])

                    if any(v is None for v in [vx, vy_h, vy_l]):
                        raise ValueError("Unknown Value")

                    # 运行 Kernel
                    # 注意：y_h/y_l 在注册表中是以 f32 存储的 bit pattern，需传给 kernel 处理
                    # 我们需要把它们当作 bits 传进去。hex_to_jax 已经把 hex 转成了 float
                    # 这里我们需要把 float 里的 bits 取出来当 integer 用
                    vy_h_bits = lax.bitcast_convert_type(vy_h, jnp.uint32)
                    vy_l_bits = lax.bitcast_convert_type(vy_l, jnp.uint32)

                    res = op_clampa_f32(vx, vy_h_bits, vy_l_bits)
                    hex_res = jax_to_hex(res)

                test_outs.append(hex_res)

                # Check
                exp_hex = str(row[exp_col]).strip()
                if hex_res.lower().replace('0x', '') == exp_hex.lower().replace('0x', ''):
                    results.append("PASS")
                else:
                    results.append(f"FAIL (Exp: {exp_hex})")

            except Exception as e:
                test_outs.append("ERR")
                results.append(f"Error: {str(e)}")

    # ==========================================
    # 分支 B: 标准处理 (x, y)
    # ==========================================
    else:
        x_col = next((c for c in df_test.columns if c.strip() == 'x'), None)
        y_col = next((c for c in df_test.columns if c.strip() == 'y'), None)
        exp_col = next((c for c in df_test.columns if 'exp_out' in c), None)
        jit_op = jax.jit(op_func) if op_func else None

        for idx, row in df_test.iterrows():
            if pd.isna(row[x_col]):
                test_outs.append("");
                results.append("");
                continue

            # --- 1. 加载 x ---
            x_raw = str(row[x_col]).strip()
            # 优先查表
            val_x = get_val(x_raw)
            # 查不到则尝试直接解析 Hex (使用文件默认 dtype)
            if val_x is None:
                val_x = hex_to_jax(x_raw, dtype)

            try:
                if y_col:  # Binary Op
                    # --- 2. 加载 y ---
                    y_raw = str(row[y_col]).strip()
                    val_y = get_val(y_raw)

                    if val_y is None:
                        # 查不到表，说明是直接的 Hex 数值
                        # 【关键修复】: 针对 remap 算子，y 是 32位控制字，强制用 uint32 解析
                        if 'remap' in filename:
                            val_y = hex_to_jax(y_raw, jnp.uint32)
                        else:
                            # 其他情况使用文件默认 dtype
                            val_y = hex_to_jax(y_raw, dtype)

                    # 运行算子
                    res = jit_op(val_x, val_y)
                else:  # Unary Op
                    res = jit_op(val_x)

                res.block_until_ready()
                hex_res = jax_to_hex(res)
                test_outs.append(hex_res)

                # --- 3. 对比结果 ---
                exp_hex = str(row[exp_col]).strip()

                # 尝试数值对比 (解决 0x1 vs 1 的问题)
                match = False
                try:
                    clean_act = hex_res.lower().replace('0x', '').replace(' ', '')
                    clean_exp = exp_hex.lower().replace('0x', '').replace(' ', '')
                    if int(clean_act, 16) == int(clean_exp, 16):
                        match = True
                except:
                    # 回退到字符串对比
                    match = (clean_act == clean_exp)

                if match:
                    results.append("PASS")
                else:
                    results.append(f"FAIL (Exp: {exp_hex})")

            except Exception as e:
                test_outs.append("ERR")
                results.append(str(e))

    # 4. Save Excel (保持原有的 Style 逻辑)
    df_test['test_out'] = test_outs
    df_test['result'] = results

    try:
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            df_def.to_excel(writer, sheet_name='Result', index=False, startrow=0)
            start_row_test = len(df_def) + 3
            df_test.to_excel(writer, sheet_name='Result', index=False, startrow=start_row_test)

            # ... (样式代码保持不变) ...
            workbook = writer.book
            sheet = writer.sheets['Result']
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            red_font = Font(color="FF0000", bold=True)

            def set_border(min_r, max_r, min_c, max_c):
                for row in sheet.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left')

            set_border(1, len(df_def) + 1, 1, len(df_def.columns))
            test_table_start = start_row_test + 1
            test_table_end = test_table_start + len(df_test)
            set_border(test_table_start, test_table_end, 1, len(df_test.columns))

            for col_idx in range(1, len(df_test.columns) + 1):
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = 30

            result_col_idx = df_test.columns.get_loc('result') + 1
            for i in range(len(df_test)):
                row_idx = test_table_start + 1 + i
                cell = sheet.cell(row=row_idx, column=result_col_idx)
                val_str = str(cell.value)
                if val_str.startswith("FAIL") or "Error" in val_str:
                    for col in range(1, len(df_test.columns) + 1):
                        sheet.cell(row=row_idx, column=col).font = red_font

        pass_cnt = results.count("PASS")
        total_cnt = len([r for r in results if r])
        logger.info(f"  -> Saved {out_path} (Pass: {pass_cnt}/{total_cnt})")

    except Exception as e:
        logger.error(f"  [Error] Save Excel failed: {e}")


def main():
    # --- TPU Check ---
    print("=" * 60)
    print("Checking Runtime Environment...")
    devices = jax.devices()
    print(f"Running on devices: {devices}")

    # 检查是否有 TPU 设备
    is_tpu = any('TPU' in str(d) for d in devices)
    if is_tpu:
        print("✅ SUCCESS: TPU detected. Running tests on TPU hardware.")
    else:
        print("⚠️ WARNING: Not running on TPU! Tests are running on CPU/GPU.")
        print("   Bitwise results (especially NaN/subnormals) may differ from hardware spec.")
    print("=" * 60 + "\n")

    input_dir = "./input_csvs_251205"
    output_dir = "./test_results_251209_cpu"

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    files = glob.glob(os.path.join(input_dir, "*.csv"))
    print(f"Found {len(files)} files.")

    for f in files:
        run_single_file(f, output_dir)


if __name__ == "__main__":
    main()